import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color


def write_excel(data, overwrite_path= None):

    dn = data['disease_name']
    defi = data['definition']
    cate = data['category']
    ce = data['cause_symptom']
    care = data['care']
    path = data['filename']

    if overwrite_path:
        path = overwrite_path

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.title = dn

    values = ['질병명', '주제', '내용', '컨텐츠속성', dn, '정의', defi, cate, dn, '원인/증상', ce, cate, dn, '예방/치료/관리', care, cate]

    count = 0
    for r in range(1, 5):
        for c in range(1, 5):
            new_cell = sheet.cell(row=r, column=c)
            new_cell.value = values[count]
            new_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            dot = Side(border_style="hair", color="000000")
            thin = Side(border_style="thin", color="000000")
            new_cell.border = Border(left=dot,right=dot,top=dot,bottom=dot)
            if r == 1:
                fg = Color("D9D9D9")
                new_cell.fill = PatternFill("solid", fgColor=fg)
            if r != 1 and c == 2:
                fg = Color("F2F2F2")
                new_cell.fill = PatternFill("solid", fgColor=fg)
            if r != 1 and c == 3:
                new_cell.alignment = Alignment(horizontal='general', vertical='center', wrapText=True)
            if c == 4:
                new_cell.border = Border(left=dot,right=thin,top=dot,bottom=dot)
            if r == 4:
                new_cell.border = Border(left=dot,right=dot,top=dot,bottom=thin)
            if c == 4 and r == 4:
                new_cell.border = Border(left=dot,right=thin,top=dot,bottom=thin)
            count += 1

    sheet.column_dimensions['A'].width = 23
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 120
    sheet.column_dimensions['D'].width = 23
    
    try:
        wb.save(path)
        return True
    except:
        return False